import sys
import cv2
import os
import glob
import matplotlib.pyplot as plt
import numpy as np
import mysql.connector
import time
import datetime as dt
import threading
import copy
import rospy
import ros_numpy 
import actionlib 
import math

from PyQt5.QtGui import QFont, QImage, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QMessageBox, QMainWindow, QCheckBox, QTableWidgetItem, QPushButton
from login import Ui_Form           # loin
from PyQt5.QtCore import Qt, QThread, pyqtSignal
# from PyQt5.QtMultimedia import QCamera
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject, QTimer
from PyQt5.QtWidgets import QFrame, QLineEdit, QButtonGroup, QTextEdit
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from gui_window import Ui_Gui_window           # gui_window
from a_few_location import Ui_Form_1                # a_few_location
from PyQt5 import QtCore
from datetime import datetime


from product import Ui_Form_2                       # product_info
from connect_database import ConnectDatabase
from openpyxl import Workbook
from pyzbar.pyzbar import decode

from move_base_msgs.msg import MoveBaseAction, MoveBaseGoal
from env.main_enviroment import Drone_Enviroment as ENV
from sensor_msgs.msg import LaserScan, Image
from geometry_msgs.msg import Twist, PoseWithCovarianceStamped
from mavros_msgs.srv import SetMode

drone_lock = True           # Co flag de lock and unlock ( cho phep drone di chuyen hoac dung)
already_locked = False
Done = True
mode_1 = False
mode_2 = False
lan = 1
t = 0
goal_y = 0


db = mysql.connector.connect(user='root', password='123456789', host='127.0.0.1', database='data_login')
mycursor = db.cursor()

class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.uic1 = Ui_Form()  # Tạo đối tượng UI
        self.uic1.setupUi(self)  # Thiết lập giao diện

        self.setWindowFlag(Qt.FramelessWindowHint)  # Thiết lập cửa sổ không có viền
        self.setAttribute(Qt.WA_TranslucentBackground)  # Thiết lập nền của cửa sổ trong suốt

        # Kết nối nút nhấn Login, Register
        self.uic1.login_button.clicked.connect(self.login)
        self.uic1.toolButton_sign_up.clicked.connect(self.register)
        self.uic1.toolButton_account.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(0))  # Nút Click Here Sign Up được ấn chuyển qua cửa sổ 2( Register)
        self.uic1.toolButton_clik_for_sign.clicked.connect(lambda: self.uic1.Background.setCurrentIndex(1))  # Nút Already Have a Account được ấn chuyển qua cửa sổ 1(Login)

        # Kết nối nút nhấn
        self.uic1.pushButton_close.clicked.connect(self.close)
        self.uic1.pushButton_close_1.clicked.connect(self.close)
        self.uic1.checkBox_show_pass.setChecked(False)  # Khởi tạo btn show_pass ban đầu = false
        self.uic1.checkBox_show_pass.stateChanged.connect(self.show_pass)  # Kiểm tra trạng thái thay đổi của btn show_pass

        lst = ['Admin', 'User']
        for x in lst:
            self.uic1.comboBox_admin.addItem(x)
        self.uic1.line_user.setFocus()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Down:
            self.uic1.line_password.setFocus()
        if event.key() == Qt.Key_Up:
            self.uic1.line_user.setFocus()
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:  # Khi nhấn nút enter
            self.login()

    def show_pass(self):
        if self.uic1.checkBox_show_pass.isChecked():
            self.uic1.line_password.setEchoMode(QLineEdit.Normal)  # Hiện password
        else:
            self.uic1.line_password.setEchoMode(QLineEdit.Password)  # Không hiện mật khẩu

    def login(self):
        try:
            self.uic1.msg = QtWidgets.QMessageBox()
            self.uic1.msg.setWindowTitle('Warning')

            username = self.uic1.line_user.text()
            password = self.uic1.line_password.text()

            print("User", username)
            print("Password", password)

            # Tạo table login trong data_login
            query = "SELECT * FROM login WHERE BINARY Username=%s AND Password=%s"
            mycursor.execute(query, (username, password))
            result = mycursor.fetchone()
            if result:
                self.close()
                global check_in
                if result[5] == "Admin":
                    check_in = "Admin"
                else:
                    check_in = "User"

                # datetime object containing current date and time
                now = datetime.now()
                # Tách ngày và giờ
                date = now.date()
                time = now.time()
                print(type(date))
                print("date and time =", date)
                query2 = "INSERT INTO history_access (Date, Time, Username, Role) VALUES (%s, %s, %s, %s)"
                mycursor.execute(query2, (date, time, result[2], result[5]))
                db.commit()

                self.main_win1 = MainWindow()
                self.main_win1.show()
            else:
                QMessageBox.information(self, "Message", "Incorrect Username or Password")
        except Exception as e:
            print("Error during login:", e)

    def register(self):
        try:
            fullname = self.uic1.line_fullname.text()
            username = self.uic1.line_user_1.text()
            email = self.uic1.line_email.text()
            password = self.uic1.line_password_1.text()
            role = self.uic1.comboBox_admin.currentText()

            query = "SELECT * FROM login WHERE Username=%s AND Password=%s AND Email=%s AND Fullname=%s AND Role=%s"
            mycursor.execute(query, (username, password, email, fullname, role))
            result = mycursor.fetchone()

            if result:
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('Warning')
                self.uic1.msg.setText("Account Already Exists")
                self.uic1.msg.exec_()
            else:
                query1 = "INSERT INTO login (Fullname, Username, Email, Password, Role) VALUES (%s, %s, %s, %s, %s)"
                mycursor.execute(query1, (fullname, username, email, password, role))
                self.uic1.line_fullname.clear()
                self.uic1.line_user_1.clear()
                self.uic1.line_email.clear()
                self.uic1.line_password_1.clear()
                db.commit()
                self.uic1.Background.setCurrentIndex(0)
                self.uic1.msg = QtWidgets.QMessageBox()
                self.uic1.msg.setWindowTitle('')
                self.uic1.msg.setText("You have successfully registered")
                self.uic1.msg.exec_()
        except Exception as e:
            print("Error during registration:", e)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.uic = Ui_Gui_window()  # Tạo đối tượng UI
        self.uic.setupUi(self)  # Thiết lập giao diện

        self.thread = {}  # Biến thread thông thường được tạo ra bằng rỗng
        self.init_button() # Khai báo nút nhấn
        self.uic.pushButton_unlock.setEnabled(False)
        self.uic.pushButton_lock.setEnabled(False)
        self.uic.textEdit_location.setEnabled(False) 
        self.uic.pushButton_home.setEnabled(False)
        self.uic.pushButton_run.setEnabled(False)
    def init_button(self):
        # Khai báo nút nhấn
        self.uic.pushButton_unlock.clicked.connect(self.unlock)
        self.uic.pushButton_lock.clicked.connect(self.lock)
        self.uic.pushButton_home.clicked.connect(self.start_home)
        self.uic.pushButton_connect_drone.clicked.connect(self.connect_drone)
        self.uic.pushButton_run.clicked.connect(self.run_drone)



        # self.uic.pushButton_log_out.clicked.connect(self.log_out)
        self.uic.pushButton_search.clicked.connect(self.search)
        self.uic.pushButton_clear.clicked.connect(self.clear)

        self.uic.pushButton_calendar_1.clicked.connect(self.calendar1)
        self.uic.pushButton_calendar_2.clicked.connect(self.calendar2)

        self.uic.pushButton_load.clicked.connect(self.load_data)
        self.uic.pushButton_report.clicked.connect(self.report)

        self.uic.comboBox_choose_location.currentIndexChanged.connect(self.on_combobox_changed)

        global check_in
        if(check_in=="Admin"):
            self.uic.history.setEnabled(True)
            self.uic.main.setEnabled(True)
        else:
            self.uic.history.setEnabled(False)

        # Khai báo Calendar 1
        self.lich1 = QtWidgets.QCalendarWidget(self)  # Khởi tạo màn hình Calendar1 khi ấn nút
        self.lich1.setWindowTitle("Calendar")
        self.lich1.move(1100, 300)
        self.lich1.resize(400,300)
        self.lich1.clicked[QtCore.QDate].connect(self.get_data1)
        self.lich1.hide()

        # Khai báo Calendar 2
        self.lich2 = QtWidgets.QCalendarWidget(self)  # Khởi tạo màn hình Calendar2 khi ấn nút
        self.lich2.setWindowTitle("Calendar")
        self.lich2.move(1100, 400)
        self.lich2.resize(400,300)
        self.lich2.clicked[QtCore.QDate].connect(self.get_data2)
        self.lich2.hide()
    def print_text(self, text):
        self.uic.textEdit_message.append(text)

    def unlock(self):
        #self.print_text()
        self.uic.pushButton_lock.setEnabled(True)
        self.uic.pushButton_unlock.setEnabled(False)
        self.uic.pushButton_home.setEnabled(True)
        self.uic.pushButton_run.setEnabled(True)
        
        global drone_lock, already_locked
        drone_lock = False
        already_locked = False 
        # self.thread[2] = capture_video(index=2)
        # self.thread[2].start()
        # self.thread[2].signal.connect(self.show_webcam)
        # self.thread[2].message_signal.connect(self.show_message)
    def lock(self):
        self.uic.pushButton_lock.setEnabled(False)
        self.uic.pushButton_unlock.setEnabled(True)
        self.uic.pushButton_home.setEnabled(False)
        self.uic.pushButton_run.setEnabled(False)
        global drone_lock, already_locked
        drone_lock = True
   
    def get_data1(self, date):
        self.uic.dateEdit_1.setDate(date)       # Lấy giá trị ngày và giờ lịch 1
        self.lich1.hide()

    def get_data2(self, date):
        self.uic.dateEdit_2.setDate(date)       # Lấy giá trị ngày và giờ lịch 2
        self.lich2.hide()

    def connect_drone(self):
        self.uic.pushButton_unlock.setEnabled(True)
        self.uic.pushButton_connect_drone.setEnabled(False)
        self.thread[1] = connect_drone(index=1)
        self.thread[1].start()
        self.thread[1].signal.connect(self.show_webcam)
        self.thread[1].message_signal.connect(self.print_text)


    def stop_capture_video(self):
        pass

    def show_webcam(self, cv_img):
        """Cập nhật label với hình ảnh mới từ opencv"""
        #cv_img = cv2.flip(cv_img, 1)  # Lật tấm hình lại để xử lý
        qt_img = self.convert_cv_qt(cv_img)
        self.uic.label_camera.setPixmap(qt_img)

    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(800, 600, QtCore.Qt.KeepAspectRatio)
        return QPixmap.fromImage(p)

    def show_message(sefl,message):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("QR Code Status")
        msg.exec_()

    def closeEvent(self, event):
        self.stop_capture_video()
        super().closeEvent(event)

    def start_home(self):
        global z, drone_lock,already_locked,Done, mode_1, mode_2
        self.uic.textEdit_message.clear()
        #self.uic.textEdit_message.append(text)
        mode_1 =False
        #mode_2 = False
        Done = False
        self.send_goal(0, 0, 0 , 1)
        z = 2
    
    def on_combobox_changed(self, index):
        if index == 1:
            self.uic.textEdit_location.setEnabled(True) 
        else:
            self.uic.textEdit_location.setEnabled(False) 
        
    def run_drone(self):
        global z, Done, location, mode_1, mode_2, capture, check
        self.uic.textEdit_message.clear()
        Done = False
        capture = False
        check = False
        x = 0
        y = 0
        z = 0
        z_or = 0

        if self.uic.comboBox_choose_location.currentIndex() == 1:       # Chon mode A Location
            mode_1 = True
            mode_2 = False
            self.uic.textEdit_location.setEnabled(True) 
            location = self.uic.textEdit_location.toPlainText()
            tach = location.split('.')
            if len(tach) == 3:
                if tach[0].isdigit() and tach[1].isdigit() and tach[2].isdigit():
                    ke = int(tach[0])
                    du = ke % 2
                    hang = int(tach[1])
                    cot = int(tach[2])
                    if hang < 5:
                        z = hang*1.6-0.8
                    else:
                        z = hang*1.6-1.1

                    if ke == 1:
                        y = 0
                    elif ke == 2 or ke == 3:
                        y = 6.6
                    elif ke == 4 or ke == 5:
                        y = 13.2
                    else:
                        y = 19.8
                    
                    x = (cot-1)*(-2.5)

                    
                    if du == 1:
                        z_or = 0.707
                    else:
                        z_or = -0.707


                    self.send_goal(x, y, z_or,0.707)
                else:
                    QMessageBox.information(self,"Wrong","Nhập sai")
            else:
                QMessageBox.information(self, "Wrong", "Nhập thiếu")
        #    db = ConnectDatabase()
        #    result = db.search_info(location=location)
        #    if result:
        #        for row in result:
        #            mystr = "Vị trí bạn đã chọn là:"
        #            print(f"{mystr}\n{row}")
        #    else:
        #        print(" Không tìm thấy thông tin vị trí đã chọn")
        
        elif self.uic.comboBox_choose_location.currentIndex() == 2: # Chon mode Auto
            mode_2 = True
            mode_1 = False
            self.uic.textEdit_location.setEnabled(False)      
            self.send_goal(0.5,0,0.707,0.707)
            z = 0.8

        

    def send_goal(self,x, y, z,w, frame_id="map"):
        global z_or_goal, x_goal, y_goal, w_goal
        # Đợi cho server sẵn sàng
        rospy.loginfo("Waiting for move_base action server...")
        client.wait_for_server()
        x_goal = x
        y_goal = y
        z_or_goal = z
        w_goal = w

        # Tạo một mục tiêu
        goal = MoveBaseGoal()
        goal.target_pose.header.frame_id = frame_id
        goal.target_pose.header.stamp = rospy.Time.now()

        # Thiết lập vị trí mục tiêu
        goal.target_pose.pose.position.x = x
        goal.target_pose.pose.position.y = y
        goal.target_pose.pose.orientation.z = z  # Hướng mục tiêu (quaternion)
        goal.target_pose.pose.orientation.w = w

        # Gửi mục tiêu
        rospy.loginfo("Sending goal: (%f, %f, %f)", x, y, z)
        client.send_goal(goal)


    def search(self):
        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:          # Chọn mode xem Account
            start = self.uic.dateEdit_1.text()
            date_str = start
            # Định dạng chuỗi ban đầu
            date_format = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj = datetime.strptime(date_str, date_format)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str = datetime_obj.strftime("%Y-%m-%d")

            end = self.uic.dateEdit_2.text()
            date_str1 = end
            # Định dạng chuỗi ban đầu
            date_format1 = "%d/%m/%Y"
            # Chuyển đổi thành đối tượng datetime
            datetime_obj1 = datetime.strptime(date_str1, date_format1)
            # Chuyển đổi thành chuỗi mới với định dạng mong muốn
            new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")

            sql = "SELECT * FROM history_access WHERE DATE(Date) BETWEEN %s AND %s"
            if db.is_connected():
                pass
            else: db
            mycursor.execute(sql, (new_date_str, new_date_str1))
            # Lấy tất cả các bản ghi từ kết quả truy vấn
            myresult = mycursor.fetchall()

            # Tạo table với số hàng và cột
            self.uic.tableWidget_account.setRowCount(len(myresult))
            self.uic.tableWidget_account.setColumnCount(4)

            # Thiết lập font chữ
            font = QFont()
            font.setBold(True)
            font.setPointSize(12)  # set font size to 12

            # Thiết lập tiêu đề, độ rộng các cột, hàng
            self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'Username', 'Role'])
            self.uic.tableWidget_account.horizontalHeader().setFont(font)         # Set font cho tiêu đề ngang
            self.uic.tableWidget_account.verticalHeader().setVisible(False)       # Ẩn phần tiêu đề dọc

            for i in range(4):
                self.uic.tableWidget_account.setColumnWidth(i, 190)
            self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
            self.uic.tableWidget_account.setFont(font)

            table_row = 0
            for row in myresult:
                date_str3 = row[0].strftime("%d/%m/%Y")
                self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
                self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[1])))
                self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[2])))
                self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
                self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[3]))
                self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)
                table_row += 1


        elif self.uic.comboBox_chose_the_mode.currentIndex() == 2:       # Chon mode xem Detect Box
           start = self.uic.dateEdit_1.text()
           date_str = start
           # Dinh dang chuoi ban dau
           date_format = "%d/%m/%Y"
           # Chuyen doi thanh doi tuong datetime
           datetime_obj = datetime.strptime(date_str, date_format)
           # Chuyen doi thanh chuoi moi voi dinh dang mong muon
           new_date_str = datetime_obj.strftime("%Y-%m-%d")


           end = self.uic.dateEdit_2.text()
           date_str1 = end
           # Dinh dang chuoi ban dau
           date_format1 = "%d/%m/%Y"
           # Chuyen doi thanh doi tuong datetime
           datetime_obj1 = datetime.strptime(date_str1, date_format1)
           # Chuyen doi thanh chuoi moi voi dinh dang mong muon
           new_date_str1 = datetime_obj1.strftime("%Y-%m-%d")
           # start_date = '2023-04-01'
           # end_date = '2023-04-26'


           sql = "SELECT * FROM detect_box WHERE DATE(Date) BETWEEN %s AND %s"
           if db.is_connected():
                pass
           else: db
           mycursor.execute(sql, (new_date_str, new_date_str1))
           # Lay tat ca cac ban ghi tu ket qua truy van
           myresult = mycursor.fetchall()


           # Tao table voi so hang va cot
           self.uic.tableWidget_account.setRowCount(len(myresult))
           self.uic.tableWidget_account.setColumnCount(4)


           # Thiet lap font chu
           font = QFont()
           font.setBold(True)
           font.setPointSize(12)  # set font size to 12


           # Thiet lap tieu de, do rong cac cot, hang
           self.uic.tableWidget_account.setHorizontalHeaderLabels(['Date', 'Time', 'QRCode', 'Location'])
           self.uic.tableWidget_account.horizontalHeader().setFont(font)
           self.uic.tableWidget_account.verticalHeader().setVisible(False)


           for i in range(4):
               self.uic.tableWidget_account.setColumnWidth(i, 190)
           self.uic.tableWidget_account.setStyleSheet("QTableWidget {border: 3px solid black;}")
           self.uic.tableWidget_account.setFont(font)
           table_row = 0
           for row in myresult:
               date_str3 = row[1].strftime("%d/%m/%Y")      # Row[1] là cột Date
               self.uic.tableWidget_account.setItem(table_row, 0, QTableWidgetItem(str(date_str3)))
               self.uic.tableWidget_account.item(table_row, 0).setTextAlignment(Qt.AlignCenter)
               self.uic.tableWidget_account.setItem(table_row, 1, QTableWidgetItem(str(row[2])))        # Row[2] là cột Time
               self.uic.tableWidget_account.item(table_row, 1).setTextAlignment(Qt.AlignCenter)
               self.uic.tableWidget_account.setItem(table_row, 2, QTableWidgetItem(str(row[3])))        # Row[3] là cột QRCode
               self.uic.tableWidget_account.item(table_row, 2).setTextAlignment(Qt.AlignCenter)
               self.uic.tableWidget_account.setItem(table_row, 3, QTableWidgetItem(row[4]))
               self.uic.tableWidget_account.item(table_row, 3).setTextAlignment(Qt.AlignCenter)


               table_row += 1

    def clear(self):
        self.uic.tableWidget_account.clear()
        self.uic.tableWidget_account.setRowCount(0)
        self.uic.tableWidget_account.setColumnCount(0)

    def calendar1(self):
        if self.lich1.isHidden():
            self.lich1.show()
        else:
            self.lich1.hide()

    def calendar2(self):
        if self.lich2.isHidden():
            self.lich2.show()
        else:
            self.lich2.hide()

    def load_data(self):
        self.product_win = MainWindow_1()
        self.product_win.show()

    def report(self):
        if self.uic.comboBox_chose_the_mode.currentIndex()==0:
            try:
                if db.is_connected():
                    pass
                else: db
                mycursor.execute("SELECT * FROM qrcodes_info")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Information of Products.xlsx")
                db.commit()

                QMessageBox.information(self,"Message"," Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 1:
            try:
                mycursor.execute("SELECT * FROM history_access")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("History Access.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()

        if self.uic.comboBox_chose_the_mode.currentIndex() == 2:
            try:
                if db.is_connected():
                    pass
                else: db
                mycursor.execute("SELECT * FROM detect_box")
                result = mycursor.fetchall()

                # Thêm lại table_name ( tại khi in ra report thì bị mất tiêu đề ID, Name, Age)
                table_name = [i[0] for i in mycursor.description]
                print((table_name))
                print(result)

                # Khởi tạo excel
                wb = Workbook()
                ws = wb.active
                ws.title = "mysql_data"
                ws.append(table_name)  # bắt đầu ghi dữ liệu lên file excel
                for row in result:
                    ws.append(row)
                wb.save("Detected Box.xlsx")
                db.commit()

                QMessageBox.information(self, "Message", " Your report has been exported successfully")
            except:
                print("something is wrong")
            db.close()
class MainWindow_1(QWidget):               # Form products_info
    def __init__(self):
        super().__init__()
        self.uic = Ui_Form_2()  # Tạo đối tượng UI
        self.uic.setupUi(self)  # Thiết lập giao diện

        # Create a database connection object
        self.db = ConnectDatabase()

        # Connect UI elements to class variables
        self.qrcode_id = self.uic.lineEdit
        # # Restrict input to integers
        # self.qrcode_id.setValidator(QIntValidator)

        self.location = self.uic.lineEdit_2
        self.type = self.uic.lineEdit_4
        self.status = self.uic.lineEdit_5
        self.date_of_import = self.uic.lineEdit_6

        self.add_btn = self.uic.add_btn
        self.update_btn = self.uic.update_btn
        self.select_btn = self.uic.select_btn
        self.search_btn = self.uic.search_btn
        self.clear_btn = self.uic.clear_btn
        self.delete_btn = self.uic.delete_btn

        self.result_table = self.uic.tableWidget
        self.result_table.setSortingEnabled(False)
        self.buttons_list = self.uic.function_frame.findChildren(QPushButton)

        # Initialize signal-slot conections
        self.init_signal_slot()

        # Populate the initial data in the table and Type of goods/ Status dropdowns
        self.search_info()
        # self.update_type_status()

    def init_signal_slot(self):
        # Connect buttons to their respective functions
        self.add_btn.clicked.connect(self.add_info)
        self.search_btn.clicked.connect(self.search_info)
        self.clear_btn.clicked.connect(self.clear_form_info)
        self.select_btn.clicked.connect(self.select_info)
        self.update_btn.clicked.connect(self.update_info)
        self.delete_btn.clicked.connect(self.delete_info)

    def search_info(self):
        # Function to search for qrcode information and populate the table
        # self.update_type_status()
        qrcode_info = self.get_qrcode_info()

        qrcode_result = self.db.search_info(
            qrcode_id=str(qrcode_info["qrcode_id"]),
            location=qrcode_info["location"],
            date_of_import=qrcode_info["date_of_import"],
            type=qrcode_info["type"],
            status=qrcode_info["status"],
        )

        self.show_data(qrcode_result)





    def add_info(self):
        # Function to add qrcode information
        self.disable_buttons()

        qrcode_info = self.get_qrcode_info()

        if qrcode_info["qrcode_id"] and qrcode_info["location"]:
            check_result = self.check_qrcode_id(qrcode_id=str(qrcode_info["qrcode_id"]))    # Kiểm tra xem đã tồn tại qrcode hay chưa

            if check_result:
                QMessageBox.information(self,"Warning","Please input a new QR Code ID",QMessageBox.StandardButton.Ok)
                self.enable_buttons()
                return

            add_result = self.db.add_info(qrcode_id=str(qrcode_info["qrcode_id"]),
                                          location=qrcode_info["location"],
                                          date_of_import=qrcode_info["date_of_import"],
                                          type=qrcode_info["type"],
                                          status=qrcode_info["status"],
                                          )
            if add_result:
                QMessageBox.information(self,"Warning",f"Add fail:{add_result}, Please try again.",QMessageBox.StandardButton.Ok)

        else:
            QMessageBox.Information(self,"Warning","Please input QR Code ID and location",QMessageBox.StandardButton.Ok)

        self.search_info()
        self.enable_buttons()
    def clear_form_info(self):
        # Function to clear the form
        # self.update_type_status()
        self.qrcode_id.clear()
        self.qrcode_id.setEnabled(True)
        self.location.clear()
        self.type.clear()
        self.status.clear()
        self.date_of_import.clear()


    def update_info(self):
        # Function to update qrcode information
        new_qrcode_info = self.get_qrcode_info()

        if new_qrcode_info["qrcode_id"]:
            update_result = self.db.update_info(
                qrcode_id=str(new_qrcode_info["qrcode_id"]),
                location=new_qrcode_info["location"],
                date_of_import=new_qrcode_info["date_of_import"],
                type=new_qrcode_info["type"],
                status=new_qrcode_info["status"],
            )

            if update_result:
                QMessageBox.information(self,"Warning", f"Fail to update the information:{update_result}, Please try again.",
                                        QMessageBox.StandardButton.Ok)
            else:
                self.search_info()

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information to update.")


    def delete_info(self):
        select_row = self.result_table.currentRow()
        if select_row != -1:
            select_option = QMessageBox.warning(self, "Warning", "Are you sure to delete it?",
                                                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
            if select_option == QMessageBox.StandardButton.Ok:
                qrcode_id = self.result_table.item(select_row, 0).text().strip()
                print(f"Deleting QR Code ID: {qrcode_id}")
                delete_result = self.db.delete_info(qrcode_id=qrcode_id)
                print(f"Delete Result: {delete_result}")
                if not delete_result:
                    self.search_info()
                else:
                    QMessageBox.information(self, "Warning",
                                            f"Fail to delete the information: {delete_result}, please try again",
                                            QMessageBox.StandardButton.Ok)
            else:
                self.search_info()
        else:
            QMessageBox.information(self, "Warning", "Please select one qrcode information to delete",
                                    QMessageBox.StandardButton.Ok)

    def select_info(self):
        # Function to select and populate student information in the form
        select_row = self.result_table.currentRow()
        if select_row != -1:      # Nếu có hàng được chọn
            self.qrcode_id.setEnabled(False)
            qrcode_id = self.result_table.item(select_row,0).text().strip()
            location = self.result_table.item(select_row, 1).text().strip()
            type = self.result_table.item(select_row, 2).text().strip()
            status = self.result_table.item(select_row, 3).text().strip()
            date_of_import = self.result_table.item(select_row, 4).text().strip()

            self.qrcode_id.setText(qrcode_id)
            self.location.setText(location)
            self.type.setText(type)
            self.status.setText(status)
            self.date_of_import.setText(date_of_import)

        else:
            QMessageBox.information(self,"Warning", "Please select one qrcode information",
                                    QMessageBox.StandardButton.Ok)


    def disable_buttons(self):
        # Disable all the buttons
        for button in self.buttons_list:
            button.setDisabled(True)

    def enable_buttons(self):
        # enable all the buttons
        for button in self.buttons_list:
            button.setDisabled(False)

    def get_qrcode_info(self):
        # Function to retrieve qrcode information from the form
        qrcode_id = self.qrcode_id.text().strip()
        location = self.location.text().strip()
        date_of_import = self.date_of_import.text().strip()
        type = self.type.text().strip()
        status = self.status.text().strip()

        qrcode_info = {
            "qrcode_id": qrcode_id,
            "location": location,
            "date_of_import": date_of_import,
            "type": type,
            "status": status,
        }

        return qrcode_info

    def check_qrcode_id(self,qrcode_id):
        # Function to check if a qrcode_id already exists
        result = self.db.search_info(qrcode_id=qrcode_id)

        return result

    def show_data(self,result):    # Hiển thị dữ liệu trong bảng kết quả
        # Function to populate the table with qrcode information
        if result:
            self.result_table.setRowCount(0)
            self.result_table.setRowCount(len(result))

            for row, info in enumerate(result):
                info_list =[
                    info["qrcode"],
                    info["location"],
                    info["type"],
                    info["status"],
                    info["date_of_import"],
                ]

                for column, item in enumerate(info_list):
                    cell_item = QTableWidgetItem(str(item))
                    self.result_table.setItem(row,column,cell_item)
        else:
            self.result_table.setRowCount(0)
            return
        


class connect_drone(QThread):
    signal = pyqtSignal(np.ndarray)         # tin hieu cv_img
    message_signal = pyqtSignal(str)        # tin hieu message


    def __init__(self, index=0):
        super(connect_drone,self).__init__()
        self.main = MainWindow() 
        self.index = index
        self.qr_codes_scanned = []
        self.qr_codes_scanned_2 = []
        self.location_qrcode = []
        self.location_total = []
        for a in range(1, 3):
            for b in range(1, 9):
                for c in range(1, 9):
                    self.location_total.append(f"{a}.{b}.{c}")
        print("Start threading connect drone", self.index)
        # super(capture_video,self).__init__()

    def run(self):
        global action, drone_lock, already_locked, high, observation, z,Done, capture
        global x, y , z_or, w_or, yaw
        global mode_1, mode_2
        global hang, ke, check
        global goal_y
        global z_or_goal, x_goal, y_goal, w_goal
        z_or_goal = 0
        x_goal = 0
        y_goal = 0
        capture = False
        enable_x = False
        pid =[1.0, 1.0, 0.0]
        pre_error_x = 0
        pre_error_y = 0
        pre_error_z = 0
        pre_error_z_or = 0
        z = 0
        high = 0
        pre_high = 0
        pre_x = 0
        pre_y = 0
        pre_z_or = 0
        pre_w_or = 1
        x = 0
        y = 0
        linear_x = 0
        linear_y = 0
        linear_z = 0
        hang = 1
        ke = 1

        observation = env.reset()
        action = np.array([[0,0,0],[0,0,0]]) 
        lidar_1d_sub = rospy.Subscriber("/lidar_1d_scan", LaserScan, self.lidar_callback)
        local_img_sub = rospy.Subscriber("/iris_rplidar/usb_cam/image_raw", Image, self.img_cb)
        cmd_vel_sub = rospy.Subscriber('/cmd_vel', Twist, self.cmd_vel_callback)
        amcl_sub = rospy.Subscriber('/amcl_pose', PoseWithCovarianceStamped, self.amcl_pose_callback)
        while True:
            # Drone is locked
            if drone_lock:
                if already_locked:
                    pass 
                else:
                    old_observation = copy.copy(observation)
                    already_locked = True
                next_observation, reward, done, info = env.position_step(old_observation.pose)
                observation = next_observation
                pre_high = high
            # Drone is unlocked
            else:
                if Done:
                    if mode_2:  #mode auto
                        if hang < 5:
                            goal_z = hang*1.6-0.8
                        else:
                            goal_z = hang*1.6-1.2
                        
                        du_h = (hang+ke) % 2 #
                        if du_h == 0:
                            goal_x = -18 # cong them 0.5
                        else:
                            goal_x = 0.5  # cong them 0.5
                        
                        if ke == 1:
                            goal_y = 0
                        elif ke == 2 or ke == 3:
                            goal_y = 6.7
                        elif ke == 4 or ke == 5:
                            goal_y = 13.4
                        else:
                            goal_y = 20.1
                        du_k = (ke+1) // 2

                        error_y = goal_y - y
                        if (error_y < -0.1) or (error_y > 0.1):
                                if -1 < error_y < 1:
                                    linear_y = pid[0]*error_y + pid[1]*(error_y - pre_error_y)
                                elif error_y > 1:
                                    if linear_y < 1:
                                        linear_y += 0.02
                                else:
                                    if linear_y > -1:
                                        linear_y -= 0.02     
                        else:
                            linear_y = 0                 

                        error_z_or = 0.707*(-1)**(ke+1) - z_or
                        if (error_z_or < -0.02) or (error_z_or > 0.02):
                            linear_z_or = pid[0]*error_z_or*1.5 + pid[1]*1.5*(error_z_or - pre_error_z_or)
                        else:
                            linear_z_or = 0

                        error_z = goal_z - high    
                        #print(pre_high,high)                
                        if (error_z < -0.1) or (error_z > 0.1):
                            action[0][2] = pid[0]*error_z*1.5 + pid[1]*(error_z - pre_error_z)
                        else:
                            action[0][2] = 0
                            enable_x = True

                        if enable_x :
                            error_x = goal_x - x
                            if (error_x < -0.1) or (error_x > 0.1):
                                capture = True
                                if -1 < error_x < 1:
                                    linear_x = pid[0]*error_x + pid[1]*(error_x - pre_error_x)
                                elif error_x > 1:
                                    if linear_x < 1:
                                        linear_x += 0.02
                                else:
                                    if linear_x > -1:
                                        linear_x -= 0.02                                    
                            else:
                                capture = False
                                linear_x = 0
                                hang = hang + (-1)**(ke+1)
                                if hang > 2 or hang < 1:  # chỉnh số lượng hàng
                                    hang = hang + (-1)**(ke)
                                    Done = False
                                    capture = False
                                    if ke == 2:   #chỉnh số lượng kệ
                                        check = True
                                        self.main.send_goal(0,0, 0, 1)
                                        z = 2                                 
                                        ke = 1
                                    else:
                                        self.main.send_goal(goal_x, du_k*6.7, (-1)**ke*0.707, 0.707)
                                        z = goal_z
                                        ke = ke + 1

                                # else:
                                #     self.main.send_goal(goal_x,goal_y, (-1)**(ke+1)*0.707, 0.707)
                                #     Done = False
                                #     capture = False
                                    
                                    

                                    
                                enable_x =False


                        action[0][0] = linear_x*math.cos(yaw)+linear_y*math.sin(yaw)
                        action[0][1] = -(linear_x*math.sin(yaw))+linear_y*math.cos(yaw)
                        action[1][2] = linear_z_or
                        next_observation, reward, done, info = env.velocity_step(action)
                        pre_error_x = error_x
                        pre_error_y = error_y
                        pre_error_z = error_z
                        pre_error_z_or = error_z_or
                    else:   #mode a location hoặc giữ vị trí
                        error_x = pre_x - x
                        if (error_x < -0.1) or (error_x > 0.1):
                            linear_x = pid[0]*error_x + pid[1]*(error_x - pre_error_x)
                        else:
                            linear_x = 0                 

                        error_y = pre_y - y
                        if (error_y < -0.1) or (error_y > 0.1):
                            linear_y = pid[0]*error_y + pid[1]*(error_y - pre_error_y)
                        else:
                            linear_y = 0

                        error_z = pre_high - high    
                        #print(pre_high,high)                
                        if (error_z < -0.1) or (error_z > 0.1):
                            action[0][2] = pid[0]*error_z*1.5 + pid[1]*(error_z - pre_error_z)
                        else:
                            action[0][2] = 0
                        
                        error_z_or = pre_z_or - z_or
                        if (error_z_or < -0.02) or (error_z_or > 0.02):
                            linear_z_or = pid[0]*error_z_or*1.5 + pid[1]*1.5*(error_z_or - pre_error_z_or)
                        else:
                            linear_z_or = 0

                        action[0][0] = linear_x*math.cos(yaw)+linear_y*math.sin(yaw)
                        action[0][1] = -(linear_x*math.sin(yaw))+linear_y*math.cos(yaw)
                        action[1][2] = linear_z_or
                        #print(action)
                        next_observation, reward, done, info = env.velocity_step(action)
                        pre_error_x = error_x
                        pre_error_y = error_y
                        pre_error_z = error_z
                        pre_error_z_or = error_z_or                   
                else:
                    if z >0:
                        pre_high = z
                    error_z = pre_high - high   
                    error_z_or = z_or_goal - z_or          
                    if (error_z < -0.1) or (error_z > 0.1):
                        action[0][2] = pid[0]*error_z*1.5 + pid[1]*(error_z - pre_error_z)
                    else:
                        action[0][2] = 0
                    # if (error_z_or > 0.02) or (error_z_or < -0.02):
                    #     action[1][2] = pid[0]*error_z_or + pid[1]*(error_z_or - pre_error_z_or)
                    #     #print(action[1][2],error_z_or)
                    # else:
                    #     action[1][2] = 0  
                        if client.get_state() == actionlib.GoalStatus.SUCCEEDED:
                                # error_x = x_goal - x
                                # if (error_x < -0.1) or (error_x > 0.1):
                                #     linear_x = pid[0]*error_x + pid[1]*(error_x - pre_error_x)
                                # else:
                                #     linear_x = 0                 

                                # error_y = y_goal - y
                                # if (error_y < -0.1) or (error_y > 0.1):
                                #     linear_y = pid[0]*error_y + pid[1]*(error_y - pre_error_y)
                                # else:
                                #     linear_y = 0          
                                # action[0][0] = linear_x*math.cos(yaw)+linear_y*math.sin(yaw)
                                # action[0][1] = -(linear_x*math.sin(yaw))+linear_y*math.cos(yaw)       
                         
                            rospy.loginfo("Goal reached!")
                            pre_x = x_goal
                            pre_y = y_goal
                            pre_z_or = z_or_goal
                            pre_w_or = w_goal
                            capture = True
                            Done = True
                            old_observation = copy.copy(observation)
                    #print(pre_high,high)
                    next_observation, reward, done, info = env.velocity_step(action)
                    observation = next_observation
                    pre_error_z = error_z
                    pre_error_z_or = error_z_or   
    def amcl_pose_callback(self,pose):
        global x, y, z_or, w_or, yaw
        x = pose.pose.pose.position.x
        y = pose.pose.pose.position.y
        z_or = pose.pose.pose.orientation.z
        w_or = pose.pose.pose.orientation.w
        yaw = math.atan2(2*(w_or*z_or ), 1-2*( z_or**2))


    def lidar_callback(self,data):
        global high
        # In các giá trị khoảng cách đầu tiên
        high = data.ranges[-1]
        #print(high)
    def cmd_vel_callback(self,msg):
        global action
        # In các giá trị khoảng cách đầu tiên
        linear_x = 1.5*msg.linear.x
        linear_y = 1.5*msg.linear.y
        angular_z = msg.angular.z
        action = np.array([[linear_x,linear_y,0],[0,0,angular_z]]) 
    def img_cb(self, data):
        global capture, location, lan,t
        global hang, ke, x,y,goal_y, check, mode_1, mode_2
        global x_local, y_local
        image = ros_numpy.numpify(data).copy()
        frame = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        # Lấy chiều cao và chiều rộng của ảnh
        height, width = frame.shape[:2]

        # Tính tọa độ tâm điểm
        center_x = width / 2


        #print("Tọa độ tâm điểm: (", center_x, ",", center_y, ")")

        if capture:
            t = t + 1
            barcodes = decode(frame)
            if barcodes:
                for barcode in barcodes:
                    # Lay toa do cua hinh chu nhat bao quanh ma QR
                    points = barcode.polygon
                    pts = [(point.x, point.y) for point in points]
                    pts = np.array(pts, np.int32)
                    pts = pts.reshape((-1, 1, 2))


                    # Neu so diem khong bang 4, bo qua ma nay
                    if len(pts) == 4:
                        #print(p)
                        barcode_data = barcode.data.decode('utf-8')
                        barcode_type = barcode.type
                        text = f"{barcode_data}"  
                        # Lay thoi gian hien tai
                        now = datetime.now()
                        # Tach ngay va gio
                        date = now.date()
                        current_time = now.time()
 



                        frame = cv2.polylines(frame, [pts], True, (0, 255, 0), 3)
                        cv2.putText(frame, text, (pts[0][0][0], pts[0][0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5,(0, 0, 255), 2, cv2.LINE_AA)

                        x_qrcode,y_qrcode,w,h = cv2.boundingRect(pts)

                        # Tính toán tâm điểm
                        x_qr_frame = x_qrcode + w / 2
                         
                        #print(pts)
                        if mode_1:
                            #Luu ma QRCode vao database
                            self.save_qrcode_to_db(date, current_time,barcode_data)

                            # Lưu mã QR Code vào danh sách các mã đã quét
                            if barcode_data not in self.qr_codes_scanned:
                                self.qr_codes_scanned.append(barcode_data)

                            if lan < 3:
                                current_time = datetime.now()
                                title = current_time.strftime("%Y-%m-%d-%Hh%M")
                                cv2.imwrite(f'/home/hai/catkin_ws/image/{location}/{title}_{lan}.jpg', frame)
                                lan = lan+1
                                print('chup_qr')
                        elif mode_2:
                            if barcode_data not in self.qr_codes_scanned_2:
                                self.qr_codes_scanned_2.append(barcode_data)
                                # du_cot = ke%2
                                # if du_cot == 1:
                                #     x_qr = x - (center_x - x_qr_frame)*((2+goal_y-y)*4.8/2)/320

                                # else:
                                #     x_qr = x + (center_x - x_qr_frame)*((2-goal_y+y)*4.2/2)/320                                
                                # if x_qr < -16:   #-16.25
                                #     cot = 8
                                # elif x_qr < -13.5:   #-13.75
                                #     cot = 7
                                # elif x_qr < -11:  #-11.25
                                #     cot = 6
                                # elif x_qr < -8.5:  #-8.75
                                #     cot = 5
                                # elif x_qr < -6.25:
                                #     cot = 4
                                # elif x_qr < -3.75:
                                #     cot = 3
                                # elif x_qr < -1.25:
                                #     cot = 2
                                # else:
                                #     cot = 1
                                
                                du_cot = (ke+hang)%2
                                if du_cot == 0:
                                    if x > -0:
                                        cot = 1
                                    elif x > -2.5:
                                        cot = 2
                                    elif x > -5:
                                        cot = 3
                                    elif x > -7.5:
                                        cot = 4
                                    elif x > -10:
                                        cot = 5
                                    elif x > -12.5:
                                        cot = 6
                                    elif x > -15:
                                        cot = 7
                                    else:
                                        cot = 8
                                else:
                                    if x < -17.5:
                                        cot = 8
                                    elif x < -15:
                                        cot = 7
                                    elif x < -12.5:
                                        cot  = 6
                                    elif x < -10:
                                        cot  = 5
                                    elif x < -7.5:
                                        cot  = 4
                                    elif x < -5:
                                        cot = 3
                                    elif x < -2.5:
                                        cot  = 2
                                    else:
                                        cot  = 1

                                location_2 = f"{ke}.{hang}.{cot}"
                                self.location_qrcode.append(location_2)
                                self.check_qrcode_mode_2(barcode_data,location_2)
                                print(location_2,x)



                    else:
                        pass
            if mode_1:
                if t > 30:
                    if len(self.qr_codes_scanned) > 0:
                        self.check_qrcode_mode_1(self.qr_codes_scanned,location)
                        self.qr_codes_scanned = []
                    else:
                        self.message_signal.emit(f"Không có hàng")
                        print("Không quét được QR Code nào")
                        current_time = datetime.now()
                        title = current_time.strftime("%Y-%m-%d-%Hh%M")
                        cv2.imwrite(f'/home/hai/catkin_ws/image/{location}/{title}_{lan}.jpg', frame)
                        print('chup')
                    capture = False
                    lan = 1
                    t = 0
        if mode_2:
            if check:
                if len(self.qr_codes_scanned_2) > 0:
                    self.check_quatity_mode_2(self.qr_codes_scanned_2,self.location_qrcode)
                    self.qr_codes_scanned_2 = []  
                    self.location_qrcode = []
                else:
                    self.message_signal.emit(f"Không quét được QR Code nào")
                    print("Không quét được QR Code nào")   
                mode_2 = False
                check = False  
                t = 0            

            time.sleep(0.1)
        self.signal.emit(frame)
    def stop(self):
       print("stop threading capture video", self.index)
       self.is_running = False
       if self.cap is not None:
           self.cap.release()      # Giai phong camera
       self.terminate()    # Dung thread
       self.wait()

    def save_qrcode_to_db(self,date,current_time,barcode_data):
       global location
       try :
       # Kiem tra QRCode da ton tai trong detect_box vao ngay hom nay hay chua
           query_check = "SELECT COUNT(*) FROM detect_box WHERE QRCode =%s AND Date =%s"
           if db.is_connected():
               pass
           else: db
           mycursor.execute(query_check,(barcode_data,date))
           result = mycursor.fetchone()


           if result[0] == 0:          # QRCode chua ton tai thi chen vao
               query_insert = "INSERT INTO detect_box (Date, Time, QRCode, Location) VALUES (%s, %s, %s, %s)"
               if db.is_connected():
                    pass
               else:
                   db
               mycursor.execute(query_insert, (date, current_time, barcode_data, location))
               db.commit()

          
           else : #QRCode da ton tai, cap nhat thoi gian moi nhat
               query_update = "UPDATE detect_box SET Time = %s WHERE QRCode = %s AND Date = %s"
               if db.is_connected():
                    pass
               else:
                   db
               mycursor.execute(query_update,(current_time,barcode_data,date))
               db.commit()



       except mysql.connector.Error as err:
           print(f"Error: {err}")

    def check_qrcode_mode_1(self,qr_code,location):
        n = len(qr_code)
        for i in range(n):
            query1 = "SELECT location FROM qrcodes_info WHERE qrcode = %s"
            if db.is_connected():
                    pass
            else: db           
            mycursor.execute(query1, (qr_code[i],))
            result = mycursor.fetchone()

            if result:
                if location == result[0]:
                    self.message_signal.emit(f"QR_Code: {qr_code[i]} - Đúng vị trí")
                else:
                    self.message_signal.emit(f"QR_Code: {qr_code[i]} - Sai vị trí, vị trí đúng: {result[0]}")
            else:
                self.message_signal.emit(f"QR_Code: {qr_code[i]} - Không có trong CSDL")
        query2 = "SELECT COUNT(*) FROM qrcodes_info WHERE location = %s"
        if db.is_connected():
            pass
        else: db
        mycursor.execute(query2, (location,))
        count = mycursor.fetchone()
        if n != count[0]:
            self.message_signal.emit(f"Sai số lượng")
        else:
            self.message_signal.emit(f"Đúng số lượng")
        self.message_signal.emit(f"Finished!")
    def check_qrcode_mode_2(self,qr_code,location):
            query1 = "SELECT location FROM qrcodes_info WHERE qrcode = %s"
            if db.is_connected():
                pass
            else: db
            mycursor.execute(query1, (qr_code,))
            result_1 = mycursor.fetchone()

            if result_1:
                if location == result_1[0]:
                    #self.message_signal.emit(f"oke")
                    pass
                else:
                    self.message_signal.emit(f"QR_Code: {qr_code} - Sai vị trí, vị trí hiện tại: {location}, vị trí đúng: {result_1[0]}")
            else:
                self.message_signal.emit(f"QR_Code: {qr_code} - Không có trong CSDL")

    def check_quatity_mode_2(self,qr_code, location):
        n = len(qr_code)
        # Chuẩn bị câu truy vấn với số lượng placeholder phù hợp
        placeholders = ', '.join(['%s'] * n)
        query2 = f"SELECT qrcode FROM qrcodes_info WHERE qrcode NOT IN ({placeholders})"
        if db.is_connected():
            pass
        else: db
        mycursor.execute(query2, qr_code)
        result_2 = mycursor.fetchall()
        if result_2:
            pass
            qr_codes_str = ', '.join([qr[0] for qr in result_2])
            self.message_signal.emit(f"Không tìm thấy QR_Code: {qr_codes_str}")
        # Chuyển đổi mảng con thành tập hợp để loại bỏ các phần tử trùng lặp
        sub_set = set(location)
        # Tạo mảng chứa các phần tử trong mảng cha không có trong mảng con
        result_array = [item for item in self.location_total if item not in sub_set]
        self.message_signal.emit(f"Vị trí trống hiện tại: {result_array}")
        self.message_signal.emit(f"Finished!")



            


if __name__ == "__main__":
    # Chạy ứng dụng
    env = ENV()
    # Tạo client cho action MoveBase
    client = actionlib.SimpleActionClient('move_base', MoveBaseAction)
    rospy.init_node("main_enviroemnt")
    app = QApplication(sys.argv)
    login_win = LoginWindow()
    login_win.show()
    sys.exit(app.exec())